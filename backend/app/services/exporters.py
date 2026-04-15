"""Human-friendly export formats for a scan.

JSON is machine-readable but brutal to skim. These exporters render the same
ScanDetail payload as a spreadsheet / self-contained HTML report / markdown
doc, optimized for "drop this in front of a non-engineer" moments.

All functions are pure: they take the ScanDetail dict (as returned by
`ScanDetail.model_dump(mode='json')`) and return bytes ready to stream.
"""

from __future__ import annotations

import csv
import html
import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

SEVERITY_COLORS = {
    "info": "#5a8dee",
    "low": "#3bc275",
    "medium": "#e0a43b",
    "high": "#e0573b",
    "critical": "#b91c1c",
}


def _flatten_data(data: Any) -> str:
    """Collapse a finding's `data` dict to a short human summary."""
    if not isinstance(data, dict) or not data:
        return ""
    parts = []
    for k, v in data.items():
        if isinstance(v, list):
            parts.append(f"{k}: [{len(v)} items]")
        elif isinstance(v, dict):
            parts.append(f"{k}: {{{len(v)} keys}}")
        elif v is None:
            continue
        else:
            s = str(v)
            parts.append(f"{k}: {s[:80]}" + ("…" if len(s) > 80 else ""))
    return " | ".join(parts)


# ---------------------------------------------------------------------- CSV --


def to_csv(scan: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(
        ["module", "severity", "title", "summary", "created_at", "data_json"]
    )
    for f in scan.get("findings", []):
        w.writerow(
            [
                f.get("module", ""),
                f.get("severity", ""),
                f.get("title", ""),
                _flatten_data(f.get("data")),
                f.get("created_at", ""),
                json.dumps(f.get("data") or {}, ensure_ascii=False),
            ]
        )
    return buf.getvalue().encode("utf-8-sig")  # BOM → Excel opens UTF-8 cleanly


# --------------------------------------------------------------------- XLSX --


def to_xlsx(scan: dict) -> bytes:
    wb = Workbook()

    # --- Overview sheet -----------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.append(["Field", "Value"])
    for row in [
        ("Domain", scan.get("domain", "")),
        ("Scan ID", scan.get("id", "")),
        ("Status", scan.get("status", "")),
        ("Started", scan.get("created_at", "")),
        ("Finished", scan.get("finished_at", "") or "—"),
        ("Findings", len(scan.get("findings", []))),
        ("Modules", len(scan.get("module_runs", []))),
    ]:
        ws.append(row)
    _style_header(ws)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60

    # --- Findings sheet -----------------------------------------------------
    fw = wb.create_sheet("Findings")
    fw.append(["Module", "Severity", "Title", "Summary", "Created", "Data"])
    for f in scan.get("findings", []):
        fw.append(
            [
                f.get("module", ""),
                f.get("severity", ""),
                f.get("title", ""),
                _flatten_data(f.get("data")),
                f.get("created_at", ""),
                json.dumps(f.get("data") or {}, ensure_ascii=False),
            ]
        )
    _style_header(fw)
    for col, width in zip("ABCDEF", [12, 10, 50, 60, 22, 80]):
        fw.column_dimensions[col].width = width

    # --- Module runs sheet --------------------------------------------------
    mw = wb.create_sheet("Modules")
    mw.append(["Module", "Status", "Started", "Finished", "Error"])
    for m in scan.get("module_runs", []):
        mw.append(
            [
                m.get("module", ""),
                m.get("status", ""),
                m.get("started_at", "") or "",
                m.get("finished_at", "") or "",
                m.get("error", "") or "",
            ]
        )
    _style_header(mw)
    for col, width in zip("ABCDE", [14, 10, 22, 22, 60]):
        mw.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _style_header(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2933")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


# ----------------------------------------------------------------- Markdown --


def to_markdown(scan: dict) -> bytes:
    out: list[str] = []
    out.append(f"# Recon report: {scan.get('domain', '')}")
    out.append("")
    out.append(f"- **Scan ID:** `{scan.get('id', '')}`")
    out.append(f"- **Status:** {scan.get('status', '')}")
    out.append(f"- **Started:** {scan.get('created_at', '')}")
    out.append(f"- **Finished:** {scan.get('finished_at', '') or '—'}")
    out.append(
        f"- **Findings:** {len(scan.get('findings', []))} across "
        f"{len(scan.get('module_runs', []))} modules"
    )
    out.append("")

    # Module runs summary
    out.append("## Modules")
    out.append("")
    out.append("| Module | Status | Started | Finished | Error |")
    out.append("|---|---|---|---|---|")
    for m in scan.get("module_runs", []):
        out.append(
            "| {module} | {status} | {started} | {finished} | {error} |".format(
                module=m.get("module", ""),
                status=m.get("status", ""),
                started=m.get("started_at", "") or "",
                finished=m.get("finished_at", "") or "",
                error=(m.get("error") or "").replace("\n", " ").replace("|", "\\|"),
            )
        )
    out.append("")

    # Findings grouped by module
    out.append("## Findings")
    out.append("")
    grouped: dict[str, list[dict]] = {}
    for f in scan.get("findings", []):
        grouped.setdefault(f.get("module", "?"), []).append(f)

    for module, items in grouped.items():
        out.append(f"### `{module}` — {len(items)} finding(s)")
        out.append("")
        for f in items:
            sev = f.get("severity", "info")
            out.append(f"**[{sev}]** {f.get('title', '')}")
            out.append("")
            data = f.get("data") or {}
            if data:
                out.append("```json")
                out.append(json.dumps(data, indent=2, ensure_ascii=False))
                out.append("```")
                out.append("")
    return "\n".join(out).encode("utf-8")


# --------------------------------------------------------------------- HTML --

_HTML_CSS = """
:root { color-scheme: light; }
* { box-sizing: border-box; }
body { font: 14px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
       color: #1a202c; max-width: 980px; margin: 32px auto; padding: 0 20px;
       background: #fafbfc; }
h1 { margin: 0 0 4px; font-size: 28px; }
h2 { margin-top: 36px; border-bottom: 1px solid #e2e8f0; padding-bottom: 6px; }
h3 { margin-top: 24px; font-size: 16px; }
.meta { color: #4a5568; font-size: 13px; margin-bottom: 24px; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 4px;
         font-size: 11px; text-transform: uppercase; font-weight: 600;
         color: #fff; letter-spacing: 0.4px; }
.summary { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 16px 0 28px; }
.summary .box { background: #fff; border: 1px solid #e2e8f0; border-radius: 6px;
                padding: 12px 14px; }
.summary .box .n { font-size: 22px; font-weight: 600; }
.summary .box .l { font-size: 11px; text-transform: uppercase;
                   letter-spacing: 0.5px; color: #718096; }
table { width: 100%; border-collapse: collapse; margin: 8px 0 16px;
        background: #fff; border: 1px solid #e2e8f0; border-radius: 6px;
        overflow: hidden; }
th, td { text-align: left; padding: 8px 12px; border-bottom: 1px solid #edf2f7;
         vertical-align: top; font-size: 13px; }
th { background: #f7fafc; font-weight: 600; font-size: 12px;
     text-transform: uppercase; letter-spacing: 0.3px; color: #4a5568; }
tr:last-child td { border-bottom: none; }
.finding { background: #fff; border: 1px solid #e2e8f0; border-radius: 6px;
           padding: 14px 16px; margin: 10px 0; }
.finding .head { display: flex; align-items: center; gap: 10px; margin-bottom: 6px; }
.finding .title { font-weight: 600; }
pre { background: #f7fafc; padding: 10px 12px; border-radius: 4px;
      font-size: 12px; overflow-x: auto; margin: 8px 0 0;
      border: 1px solid #edf2f7; }
code { background: #edf2f7; padding: 1px 5px; border-radius: 3px; font-size: 12px; }
.footer { margin-top: 48px; color: #a0aec0; font-size: 12px; text-align: center; }
@media print {
  body { background: #fff; max-width: none; margin: 0; }
  .finding, .summary .box, table { break-inside: avoid; }
}
"""


def _badge(severity: str) -> str:
    color = SEVERITY_COLORS.get(severity, "#718096")
    return f'<span class="badge" style="background:{color}">{html.escape(severity)}</span>'


def _data_as_html(data: Any) -> str:
    """Render a finding's data dict — prefer tables for list-shaped values."""
    if not isinstance(data, dict) or not data:
        return ""
    rows = []
    for k, v in data.items():
        if isinstance(v, list) and v and all(isinstance(x, str) for x in v):
            if len(v) > 20:
                body = ", ".join(html.escape(x) for x in v[:20])
                body += f' <em>…and {len(v) - 20} more</em>'
            else:
                body = ", ".join(html.escape(x) for x in v)
        elif isinstance(v, list) and v and all(isinstance(x, dict) for x in v):
            body = _list_of_dicts_table(v)
        elif isinstance(v, dict):
            body = _dict_table(v)
        elif v is None:
            body = "<em>none</em>"
        else:
            body = html.escape(str(v))
        rows.append(f"<tr><th>{html.escape(k)}</th><td>{body}</td></tr>")
    return f"<table>{''.join(rows)}</table>"


def _list_of_dicts_table(items: list[dict]) -> str:
    keys: list[str] = []
    for it in items[:50]:
        for k in it:
            if k not in keys:
                keys.append(k)
    head = "".join(f"<th>{html.escape(k)}</th>" for k in keys)
    body = ""
    for it in items[:50]:
        cells = "".join(
            f"<td>{html.escape(str(it.get(k, '')))}</td>" for k in keys
        )
        body += f"<tr>{cells}</tr>"
    extra = ""
    if len(items) > 50:
        extra = f'<tr><td colspan="{len(keys)}"><em>…and {len(items) - 50} more</em></td></tr>'
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}{extra}</tbody></table>"


def _dict_table(d: dict) -> str:
    rows = "".join(
        f"<tr><th>{html.escape(str(k))}</th><td>{html.escape(str(v))}</td></tr>"
        for k, v in d.items()
    )
    return f"<table>{rows}</table>"


def to_html(scan: dict) -> bytes:
    findings = scan.get("findings", [])
    module_runs = scan.get("module_runs", [])
    by_sev: dict[str, int] = {}
    for f in findings:
        s = f.get("severity", "info")
        by_sev[s] = by_sev.get(s, 0) + 1

    modules_rows = "".join(
        f"<tr><td><code>{html.escape(m.get('module', ''))}</code></td>"
        f"<td>{_badge(m.get('status', ''))}</td>"
        f"<td>{html.escape(m.get('started_at', '') or '')}</td>"
        f"<td>{html.escape(m.get('finished_at', '') or '')}</td>"
        f"<td>{html.escape(m.get('error', '') or '')}</td></tr>"
        for m in module_runs
    )

    grouped: dict[str, list[dict]] = {}
    for f in findings:
        grouped.setdefault(f.get("module", "?"), []).append(f)

    findings_html = []
    for module, items in grouped.items():
        findings_html.append(
            f"<h3><code>{html.escape(module)}</code> — {len(items)} finding(s)</h3>"
        )
        for f in items:
            findings_html.append(
                f'<div class="finding">'
                f'<div class="head">{_badge(f.get("severity", "info"))}'
                f'<span class="title">{html.escape(f.get("title", ""))}</span></div>'
                f'{_data_as_html(f.get("data"))}'
                f"</div>"
            )

    summary_boxes = "".join(
        f'<div class="box"><div class="n" style="color:{SEVERITY_COLORS.get(sev, "#333")}">'
        f"{by_sev.get(sev, 0)}</div><div class=\"l\">{sev}</div></div>"
        for sev in ("info", "low", "medium", "high", "critical")
    )

    body = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<title>Recon report — {html.escape(scan.get('domain', ''))}</title>
<style>{_HTML_CSS}</style></head><body>
<h1>Recon report: {html.escape(scan.get('domain', ''))}</h1>
<div class="meta">
  Scan <code>{html.escape(scan.get('id', ''))}</code> ·
  status <strong>{html.escape(scan.get('status', ''))}</strong> ·
  started {html.escape(scan.get('created_at', ''))}
  {"· finished " + html.escape(scan.get('finished_at', '')) if scan.get('finished_at') else ""}
</div>

<div class="summary">{summary_boxes}
  <div class="box"><div class="n">{len(findings)}</div><div class="l">total findings</div></div>
</div>

<h2>Modules</h2>
<table><thead><tr><th>Module</th><th>Status</th><th>Started</th><th>Finished</th><th>Error</th></tr></thead>
<tbody>{modules_rows}</tbody></table>

<h2>Findings</h2>
{"".join(findings_html) if findings_html else "<p><em>No findings.</em></p>"}

<div class="footer">Generated by recon-dashboard-yh — print this page with your browser to save as PDF.</div>
</body></html>
"""
    return body.encode("utf-8")
